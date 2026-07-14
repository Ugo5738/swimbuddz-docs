from datetime import date, datetime, time, timedelta
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


OUT_DIR = Path("outputs/july_2026_revenue_sprint_restrategy")
OUT_FILE = OUT_DIR / "daniel_july_12_31_2026_revenue_restrategy.xlsx"


COLORS = {
    "navy": "0F172A",
    "blue": "0E7490",
    "teal": "0F766E",
    "green": "15803D",
    "orange": "C2410C",
    "gray": "475569",
    "light_blue": "D9EEF7",
    "light_green": "DCFCE7",
    "light_orange": "FFEDD5",
    "light_gray": "F1F5F9",
    "white": "FFFFFF",
}


thin = Side(style="thin", color="CBD5E1")
border = Border(left=thin, right=thin, top=thin, bottom=thin)


def style_title(ws, title, subtitle=None, end_col=10):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)
    c = ws.cell(1, 1, title)
    c.font = Font(bold=True, size=18, color=COLORS["white"])
    c.fill = PatternFill("solid", fgColor=COLORS["navy"])
    c.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 28
    if subtitle:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=end_col)
        s = ws.cell(2, 1, subtitle)
        s.font = Font(italic=True, color=COLORS["gray"])
        s.fill = PatternFill("solid", fgColor=COLORS["light_gray"])
        ws.row_dimensions[2].height = 22


def style_header_row(ws, row, start_col=1, end_col=None, fill=COLORS["blue"]):
    if end_col is None:
        end_col = ws.max_column
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row, col)
        cell.font = Font(bold=True, color=COLORS["white"])
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = border


def style_table(ws, min_row, max_row, min_col, max_col):
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    for r in range(min_row + 1, max_row + 1):
        fill = COLORS["white"] if r % 2 else COLORS["light_blue"]
        for c in range(min_col, max_col + 1):
            ws.cell(r, c).fill = PatternFill("solid", fgColor=fill)


def set_widths(ws, widths):
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def add_list_validation(ws, cell_range, values):
    dv = DataValidation(type="list", formula1=f'"{",".join(values)}"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(cell_range)


def add_bool_validation(ws, cell_range):
    add_list_validation(ws, cell_range, ["TRUE", "FALSE"])


def daily_rows():
    rows = []

    def add(d, start, end, track, priority, task, steps, output, metric, asset, status="Not Started"):
        rows.append(
            {
                "Date": d,
                "Day": d.strftime("%a"),
                "Start Time": start,
                "End Time": end,
                "Track": track,
                "Priority": priority,
                "Task": task,
                "Action Steps": steps,
                "Expected Output": output,
                "Success Metric": metric,
                "Prospect / Asset": asset,
                "Status": status,
                "Create Reminder": "TRUE",
                "Reminder Lead Mins": 15,
                "Calendar Event ID": "",
                "Notes": "",
            }
        )

    add(
        date(2026, 7, 12),
        "08:30",
        "09:15",
        "Strategy",
        "P0",
        "Lock July revenue rule",
        "Write the rule: $2k by Jul 31 from either one implementation sprint or 4-5 paid audits. Confirm the three tracks and the gate dates.",
        "One clear July rule",
        "Rule written and visible on Dashboard",
        "Dashboard",
    )
    add(
        date(2026, 7, 12),
        "11:30",
        "13:00",
        "Pipeline",
        "P0",
        "Seed 40-prospect universe",
        "Add 10 warm founder/operator contacts, 15 swim/sports academy prospects, 10 agency/operator prospects, and 5 local backup buyers.",
        "40 pipeline rows",
        "40 prospects added with pain clues",
        "Pipeline Tracker",
    )
    add(
        date(2026, 7, 12),
        "16:00",
        "17:30",
        "Offers",
        "P0",
        "Prepare offer assets",
        "Finalize warm ask, swim-school lead-to-enrollment audit, agency onboarding/reporting audit, follow-up, and paid-audit close scripts.",
        "Outreach pack ready",
        "5 reusable scripts ready",
        "Offers & Templates",
    )
    add(
        date(2026, 7, 12),
        "19:30",
        "21:00",
        "Warm Network",
        "P0",
        "Send first direct asks",
        "Send 10 warm/referral messages and 5 personalized swim/sports academy messages. Record every send in Pipeline Tracker.",
        "15 messages sent",
        "15 logged touches",
        "Warm contacts + swim prospects",
    )
    add(
        date(2026, 7, 12),
        "21:00",
        "21:20",
        "Admin/Review",
        "P0",
        "Update daily metrics",
        "Record messages sent, replies, calls booked, cash collected, and one market signal.",
        "Day 1 metrics",
        "Daily Metrics row updated",
        "Daily Metrics",
    )

    custom = {
        date(2026, 7, 13): [
            ("07:30", "08:00", "Admin/Review", "P0", "Morning score and reply triage", "Check WhatsApp, LinkedIn, email, and calendar. Pick top 10 people to move today.", "Priority list", "Top 10 next actions selected", "Daily Metrics"),
            ("12:30", "13:15", "Warm Network", "P0", "Send 15 warm/referral asks", "Ask for one founder/operator intro or one painful workflow worth auditing. Keep it specific and direct.", "15 warm messages", "15 sent and logged", "Warm contacts"),
            ("15:30", "17:00", "Swim/Sports Niche", "P0", "Record 2 Loom teardown drafts", "Pick 2 swim/sports prospects with visible inquiry/trial/booking leaks. Record short teardown or write detailed notes if Loom is not ready.", "2 teardown assets", "2 personalized observations ready", "Swim prospects"),
            ("19:30", "21:00", "Agency/Operator", "P1", "Send 10 agency/operator messages", "Target agencies or operators with onboarding, reporting, intake, or dashboard pain. Offer a 48h workflow audit.", "10 messages sent", "10 sent and logged", "Agency prospects"),
        ],
        date(2026, 7, 14): [
            ("07:30", "08:15", "Proof", "P0", "Draft SwimBuddz proof one-pager", "Summarize SwimBuddz OS, cohort registration, club subscription, payment/referral asks, and workflow complexity as proof of execution.", "Proof one-pager draft", "One page drafted", "SwimBuddz proof"),
            ("12:30", "13:15", "Warm Network", "P0", "Follow up and ask for calls", "Follow up all replies and ask for 15-minute calls. Send 10 additional warm asks if reply volume is low.", "Calls requested", "3 calls requested/booked", "Warm contacts"),
            ("15:30", "17:30", "Calls/Discovery", "P0", "Discovery/call block", "Hold booked calls or send call invitations with 2 time windows. Push toward paid audit when pain is concrete.", "Calls advanced", "2 calls booked or held", "Pipeline Tracker"),
            ("19:30", "21:00", "Swim/Sports Niche", "P0", "Send 12 swim/sports messages", "Use the lead-to-enrollment leak angle. Personalize with one clue: trial, assessment, class placement, waitlist, waiver, payment, or follow-up.", "12 niche messages", "12 sent and logged", "Swim/sports prospects"),
        ],
        date(2026, 7, 15): [
            ("07:30", "08:15", "Proof", "P0", "Draft tech-gig offer page", "Write a one-page offer for AI/backend workflow implementation: problem, deliverables, timeline, price range, proof, and call CTA.", "Tech offer page", "One offer page drafted", "Offer asset"),
            ("12:30", "13:15", "Warm Network", "P0", "Send 15 high-trust touches", "Prioritize people who already know your work. Ask for specific intros to founders/operators with messy operations.", "15 touches", "15 sent and logged", "Warm contacts"),
            ("15:30", "17:30", "Calls/Discovery", "P0", "Discovery and paid-audit asks", "For every serious conversation, ask: what manual workflow costs time or money weekly? Offer the 48h paid audit.", "Audit asks", "2 paid-audit asks made", "Pipeline Tracker"),
            ("19:30", "21:00", "Agency/Operator", "P1", "Agency workflow audit push", "Send 10 agency/operator messages focused on client onboarding, reporting, intake, dashboard, and recurring follow-up workflows.", "10 agency messages", "10 sent and logged", "Agency prospects"),
        ],
        date(2026, 7, 16): [
            ("07:30", "08:15", "Strategy", "P0", "Run swim niche gate review", "Count swim/sports replies, calls booked, and paid-audit interest. Keep only if there are 5 replies, 2 calls, or 1 paid-audit signal.", "Gate decision", "Gate marked Hit/Miss", "Gate Logic"),
            ("12:30", "13:15", "Follow-up", "P0", "Follow up all non-replies", "Send concise follow-up to all prospects touched since Jul 12. Ask a specific yes/no question.", "Follow-up wave", "25 follow-ups logged", "Pipeline Tracker"),
            ("15:30", "17:30", "Calls/Discovery", "P0", "Convert replies into calls", "Push warm and interested cold replies into call slots. Use the paid-audit close script.", "Call block filled", "2 calls booked or held", "Pipeline Tracker"),
            ("19:30", "21:00", "Strategy", "P0", "Reallocate time based on gate", "If swim gate misses, shift 70 percent of time to warm network and agency/operator offers. If it hits, keep swim as 40 percent of outreach.", "Time split updated", "Next 7 days adjusted", "Dashboard"),
        ],
    }

    start = date(2026, 7, 13)
    end = date(2026, 7, 31)
    d = start
    while d <= end:
        if d in custom:
            for item in custom[d]:
                add(d, *item)
        elif d.weekday() < 5:
            add(
                d,
                "07:30",
                "08:00",
                "Admin/Review",
                "P0",
                "Morning pipeline review",
                "Review replies, due follow-ups, calls, and yesterday's market signal. Pick the 3 outcomes that matter today.",
                "Daily priority list",
                "3 outcomes selected",
                "Daily Metrics",
            )
            if d <= date(2026, 7, 24):
                add(
                    d,
                    "12:30",
                    "13:15",
                    "Warm Network",
                    "P0",
                    "Warm direct asks and follow-ups",
                    "Send 10-15 warm/referral touches or follow-ups. Ask for a call, a referral, or permission to audit one workflow.",
                    "10-15 warm touches",
                    "Touches logged",
                    "Warm contacts",
                )
            else:
                add(
                    d,
                    "12:30",
                    "13:15",
                    "Closing",
                    "P0",
                    "Close open proposals",
                    "Follow up all open proposals, ask for a decision, and offer a smaller paid audit if a sprint is too large.",
                    "Proposal decisions",
                    "All open deals touched",
                    "Pipeline Tracker",
                )
            add(
                d,
                "15:30",
                "17:30",
                "Calls/Discovery",
                "P0",
                "Call/proposal block",
                "Hold calls, book calls, or write proposals. Move serious prospects to a paid audit or implementation sprint.",
                "Pipeline advanced",
                "2 calls/proposals advanced",
                "Pipeline Tracker",
            )
            if d <= date(2026, 7, 20):
                add(
                    d,
                    "19:30",
                    "21:00",
                    "Swim/Sports Niche",
                    "P1",
                    "Swim/sports market test outreach",
                    "Send 8-12 targeted messages or 2 personalized teardowns. Only continue heavy effort if the Jul 16 gate was hit.",
                    "Niche test touches",
                    "8-12 sent or 2 teardowns ready",
                    "Swim/sports prospects",
                )
            elif d <= date(2026, 7, 24):
                add(
                    d,
                    "19:30",
                    "21:00",
                    "Agency/Operator",
                    "P1",
                    "Agency/operator conversion push",
                    "Send targeted audit offers to agencies/operators and use the proof pages to establish credibility quickly.",
                    "Agency pipeline advanced",
                    "10 touches or 1 proposal",
                    "Agency prospects",
                )
            else:
                add(
                    d,
                    "19:30",
                    "21:00",
                    "Delivery/Closing",
                    "P0",
                    "Deliver paid audit or final close work",
                    "Finish paid audit, send Loom, send implementation plan, or make final decision asks.",
                    "Delivery/close output",
                    "1 deliverable or decision ask",
                    "Active deals",
                )
            add(
                d,
                "21:00",
                "21:20",
                "Admin/Review",
                "P0",
                "Update metrics and market signal",
                "Record sends, replies, calls, proposals, cash collected, and what the market said back.",
                "Metrics updated",
                "Daily Metrics row updated",
                "Daily Metrics",
            )
        else:
            add(
                d,
                "09:00",
                "10:00",
                "Admin/Review",
                "P0",
                "Weekend pipeline cleanup",
                "Clean stages, mark dead leads, prioritize top 15, and prepare Monday follow-ups.",
                "Clean pipeline",
                "Top 15 ranked",
                "Pipeline Tracker",
            )
            add(
                d,
                "12:00",
                "13:00",
                "Follow-up",
                "P0",
                "Weekend follow-ups",
                "Send lighter follow-ups and referral asks. Avoid long cold pitches.",
                "Follow-ups sent",
                "10 follow-ups logged",
                "Pipeline Tracker",
            )
            add(
                d,
                "16:00",
                "17:30",
                "Proof/Delivery",
                "P1",
                "Proof asset or paid audit delivery",
                "Create one proof asset, Loom teardown, audit map, or implementation proposal that can be reused.",
                "Reusable asset",
                "1 asset finished",
                "Proof assets",
            )
            add(
                d,
                "19:30",
                "20:30",
                "Admin/Review",
                "P0",
                "Weekly scorecard",
                "Score messages, replies, calls, proposals, cash, and next-week allocation. Cut weak channels fast.",
                "Scorecard",
                "Allocation updated",
                "Daily Metrics",
            )
        d += timedelta(days=1)

    return rows


APP_SCRIPT = r"""const PLAN_SHEET_NAME = 'Daily Plan';
const CONFIG_SHEET_NAME = 'Reminder Config';

function syncPlanReminders() {
  const ss = SpreadsheetApp.getActive();
  const plan = ss.getSheetByName(PLAN_SHEET_NAME);
  const config = readConfig_();
  const calendar = getCalendar_(config);
  const maxEvents = Number(config['Max Events Per Run'] || 20);
  const dryRun = String(config['Dry Run'] || 'FALSE').toUpperCase() === 'TRUE';
  const prefix = config['Event Prefix'] || 'Revenue Sprint';

  const values = plan.getDataRange().getValues();
  const h = headerMap_(values[3]);
  let changed = 0;

  for (let r = 4; r < values.length; r++) {
    if (changed >= maxEvents) break;
    const row = values[r];
    if (!row[h['Date']] || String(row[h['Create Reminder']]).toUpperCase() !== 'TRUE') continue;

    const status = String(row[h['Status']] || '').trim();
    if (['Done', 'Skipped'].includes(status)) continue;

    const start = parseTime_(row[h['Date']], row[h['Start Time']]);
    const end = parseTime_(row[h['Date']], row[h['End Time']]);
    if (!start || !end || end <= start) continue;

    const title = `${prefix}: ${row[h['Task']]}`;
    const description = eventDescription_(row, h);
    const eventIdCol = h['Calendar Event ID'] + 1;
    let eventId = row[h['Calendar Event ID']];
    let event = eventId ? calendar.getEventById(eventId) : null;

    if (dryRun) {
      Logger.log(`[DRY RUN] ${title} ${start} - ${end}`);
      continue;
    }

    if (event) {
      event.setTitle(title);
      event.setTime(start, end);
      event.setDescription(description);
      resetReminder_(event, row[h['Reminder Lead Mins']]);
    } else {
      event = calendar.createEvent(title, start, end, { description });
      resetReminder_(event, row[h['Reminder Lead Mins']]);
      plan.getRange(r + 1, eventIdCol).setValue(event.getId());
    }

    changed++;
    Utilities.sleep(250);
  }

  Logger.log(`Processed ${changed} reminders. If you need more, run syncPlanReminders again after a few minutes.`);
}

function clearCalendarEventIdsOnly() {
  const ss = SpreadsheetApp.getActive();
  const plan = ss.getSheetByName(PLAN_SHEET_NAME);
  const values = plan.getDataRange().getValues();
  const h = headerMap_(values[3]);
  plan.getRange(5, h['Calendar Event ID'] + 1, values.length - 4, 1).clearContent();
}

function readConfig_() {
  const sheet = SpreadsheetApp.getActive().getSheetByName(CONFIG_SHEET_NAME);
  const rows = sheet.getDataRange().getValues();
  const config = {};
  rows.forEach(row => {
    if (row[0]) config[String(row[0]).trim()] = row[1];
  });
  return config;
}

function getCalendar_(config) {
  const id = config['Calendar ID'];
  if (id && id !== 'primary') return CalendarApp.getCalendarById(id);
  return CalendarApp.getDefaultCalendar();
}

function headerMap_(headerRow) {
  const h = {};
  headerRow.forEach((name, idx) => {
    if (name) h[String(name).trim()] = idx;
  });
  return h;
}

function parseTime_(dateValue, timeText) {
  const d = new Date(dateValue);
  const parts = String(timeText).trim().split(':').map(Number);
  d.setHours(parts[0] || 0, parts[1] || 0, 0, 0);
  return d;
}

function eventDescription_(row, h) {
  return [
    `Track: ${row[h['Track']]}`,
    `Priority: ${row[h['Priority']]}`,
    '',
    'Action steps:',
    row[h['Action Steps']],
    '',
    `Expected output: ${row[h['Expected Output']]}`,
    `Success metric: ${row[h['Success Metric']]}`,
    `Asset/prospect: ${row[h['Prospect / Asset']]}`,
    '',
    `Sheet status: ${row[h['Status']]}`
  ].join('\n');
}

function resetReminder_(event, leadMins) {
  event.removeAllReminders();
  const mins = Number(leadMins || 15);
  event.addPopupReminder(mins);
}
"""


def build_workbook():
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard"

    style_title(
        ws,
        "July 12-31, 2026 Revenue Sprint - Restrategized",
        "Goal: collect $2,000 by July 31 without betting the whole month on one cold channel.",
        8,
    )

    dashboard_rows = [
        ("Metric", "Value", "How it is calculated / why it matters"),
        ("Revenue target", 2000, "Cash collected by Jul 31, 2026."),
        ("Cash collected", "=SUM('Daily Metrics'!L:L)", "Update Daily Metrics every night."),
        ("Remaining", "=B5-B6", "Amount still required."),
        ("Weighted pipeline", "=SUM('Pipeline Tracker'!N:N)", "Amount x probability for active deals."),
        ("Replies / conversations", "=SUM('Daily Metrics'!F:F)", "Daily market feedback, not vanity outreach."),
        ("Calls booked", "=SUM('Daily Metrics'!G:G)", "A real buying process starts here."),
        ("Proposals sent", "=SUM('Daily Metrics'!I:I)", "Formal paid opportunity count."),
        ("Paid audits sold", "=SUM('Daily Metrics'!J:J)", "Fastest smaller ticket."),
        ("Sprint deposits", "=SUM('Daily Metrics'!K:K)", "Best path to full $2k."),
        ("Task completion", '=IFERROR(COUNTIF(\'Daily Plan\'!L:L,"Done")/COUNTIF(\'Daily Plan\'!A:A,">0"),0)', "Execution check."),
    ]
    for r_idx, row in enumerate(dashboard_rows, start=4):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(r_idx, c_idx, value)
    style_header_row(ws, 4, 1, 3, COLORS["blue"])
    style_table(ws, 4, 14, 1, 3)
    ws["B5"].number_format = '$#,##0'
    ws["B6"].number_format = '$#,##0'
    ws["B7"].number_format = '$#,##0'
    ws["B8"].number_format = '$#,##0'
    ws["B14"].number_format = "0%"

    track_header_row = 16
    ws.cell(track_header_row, 1, "Priority")
    ws.cell(track_header_row, 2, "Track")
    ws.cell(track_header_row, 3, "July Role")
    ws.cell(track_header_row, 4, "Reasoning")
    ws.cell(track_header_row, 5, "Offer")
    style_header_row(ws, track_header_row, 1, 5, COLORS["teal"])
    tracks = [
        ("1", "Warm network tech/ops gigs", "Primary cash path", "Highest short-term trust. Use people who know your work to reach founders/operators with manual workflow pain.", "$300-$500 audit or $1.5k-$2.5k sprint"),
        ("2", "Agency/operator automation", "Secondary cash path", "Agencies understand onboarding/reporting pain and can justify a fixed-scope workflow sprint quickly.", "$400 audit or $1.5k-$2k implementation"),
        ("3", "Western swim/sports academy", "Niche beachhead test", "Good fit with SwimBuddz proof, but cold trust is slower. Test hard for 5 days before overcommitting.", "$300-$500 Lead-to-Enrollment Leak Audit"),
        ("4", "StrokeLab digital product", "Proof and lead magnet", "Useful for authority and swim niche credibility, but not the best standalone July cash bet.", "Free/low-ticket proof asset"),
        ("5", "Corporate wellness / local SwimBuddz", "Backup channel", "Useful where relationships already exist. Do not let it distract from cash conversations.", "Corporate intro or cohort conversion"),
    ]
    for idx, row in enumerate(tracks, start=track_header_row + 1):
        for col, value in enumerate(row, start=1):
            ws.cell(idx, col, value)
    style_table(ws, track_header_row, track_header_row + len(tracks), 1, 5)

    gate_row = 24
    gates = [
        ("Gate Date", "Required Signal", "Decision"),
        ("2026-07-16", "Swim/sports has 5 replies OR 2 calls OR 1 paid-audit interest", "If missed, reduce swim cold outreach to 20 percent and shift to warm/agency."),
        ("2026-07-20", "At least 1 proposal sent OR $300+ paid audit sold", "If missed, make direct high-trust asks and offer smaller paid audit."),
        ("2026-07-24", "$500+ collected OR signed sprint/proposal", "If missed, stop broad outreach and pursue only strongest 10 leads."),
        ("2026-07-28", "$1,000+ collected or committed", "Push implementation deposits, urgent audit delivery, and referrals."),
    ]
    for r_idx, row in enumerate(gates, start=gate_row):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(r_idx, c_idx, value)
    style_header_row(ws, gate_row, 1, 3, COLORS["orange"])
    style_table(ws, gate_row, gate_row + len(gates) - 1, 1, 3)
    set_widths(ws, {"A": 22, "B": 22, "C": 58, "D": 56, "E": 38})
    ws.freeze_panes = "A4"

    # Daily Plan
    plan = wb.create_sheet("Daily Plan")
    style_title(
        plan,
        "Day-to-Day Execution Plan: Jul 12-31, 2026",
        "Use Status daily. Apps Script can create reminders from rows where Create Reminder is TRUE.",
        16,
    )
    headers = [
        "Date",
        "Day",
        "Start Time",
        "End Time",
        "Track",
        "Priority",
        "Task",
        "Action Steps",
        "Expected Output",
        "Success Metric",
        "Prospect / Asset",
        "Status",
        "Create Reminder",
        "Reminder Lead Mins",
        "Calendar Event ID",
        "Notes",
    ]
    for c_idx, header in enumerate(headers, start=1):
        plan.cell(4, c_idx, header)
    style_header_row(plan, 4, 1, len(headers), COLORS["blue"])
    rows = daily_rows()
    for r_idx, row in enumerate(rows, start=5):
        for c_idx, header in enumerate(headers, start=1):
            value = row[header]
            plan.cell(r_idx, c_idx, value)
        plan.cell(r_idx, 1).number_format = "yyyy-mm-dd"
    style_table(plan, 4, 4 + len(rows), 1, len(headers))
    add_list_validation(plan, f"L5:L{4 + len(rows)}", ["Not Started", "In Progress", "Waiting", "Done", "Skipped"])
    add_bool_validation(plan, f"M5:M{4 + len(rows)}")
    set_widths(
        plan,
        {
            "A": 12,
            "B": 8,
            "C": 11,
            "D": 11,
            "E": 20,
            "F": 9,
            "G": 28,
            "H": 70,
            "I": 30,
            "J": 32,
            "K": 24,
            "L": 15,
            "M": 16,
            "N": 17,
            "O": 34,
            "P": 34,
        },
    )
    plan.freeze_panes = "A5"
    plan.auto_filter.ref = f"A4:P{4 + len(rows)}"

    # Daily Metrics
    metrics = wb.create_sheet("Daily Metrics")
    style_title(metrics, "Daily Metrics", "Answer every night: what did the market say back?", 15)
    m_headers = [
        "Date",
        "Day",
        "Warm Messages",
        "Swim/Academy Messages",
        "Agency/Operator Messages",
        "Replies",
        "Calls Booked",
        "Calls Held",
        "Proposals Sent",
        "Paid Audits Sold",
        "Sprint Deposits",
        "Cash Collected",
        "Market Signal",
        "Next Adjustment",
        "Done?",
    ]
    for c_idx, header in enumerate(m_headers, start=1):
        metrics.cell(4, c_idx, header)
    style_header_row(metrics, 4, 1, len(m_headers), COLORS["teal"])
    d = date(2026, 7, 12)
    r = 5
    while d <= date(2026, 7, 31):
        metrics.cell(r, 1, d)
        metrics.cell(r, 2, d.strftime("%a"))
        metrics.cell(r, 15, "FALSE")
        metrics.cell(r, 1).number_format = "yyyy-mm-dd"
        r += 1
        d += timedelta(days=1)
    style_table(metrics, 4, r - 1, 1, len(m_headers))
    add_bool_validation(metrics, f"O5:O{r - 1}")
    set_widths(
        metrics,
        {
            "A": 12,
            "B": 8,
            "C": 15,
            "D": 22,
            "E": 24,
            "F": 10,
            "G": 13,
            "H": 12,
            "I": 14,
            "J": 16,
            "K": 15,
            "L": 14,
            "M": 52,
            "N": 52,
            "O": 9,
        },
    )
    metrics.freeze_panes = "A5"

    # Pipeline Tracker
    pipe = wb.create_sheet("Pipeline Tracker")
    style_title(
        pipe,
        "Pipeline Tracker",
        "Do not count outreach as progress until it creates replies, calls, proposals, cash, or useful market feedback.",
        18,
    )
    p_headers = [
        "Date Added",
        "Track",
        "Segment",
        "Market",
        "Prospect",
        "Contact",
        "URL / LinkedIn",
        "Pain Clue",
        "Current Tools",
        "Offer",
        "Stage",
        "Amount",
        "Probability",
        "Weighted Value",
        "Next Action",
        "Next Action Date",
        "Last Touch",
        "Notes",
    ]
    for c_idx, header in enumerate(p_headers, start=1):
        pipe.cell(4, c_idx, header)
    style_header_row(pipe, 4, 1, len(p_headers), COLORS["blue"])
    stages = ["Not Contacted", "Messaged", "Replied", "Call Booked", "Call Held", "Proposal Sent", "Paid Audit", "Sprint Won", "Lost", "Nurture"]
    tracks_list = ["Warm Network", "Agency/Operator", "Swim/Sports Niche", "Corporate Wellness", "StrokeLab", "Other"]
    offers = ["Paid Audit", "Implementation Sprint", "Referral Ask", "Corporate Intro", "Digital Product", "Other"]
    segment_plan = [
        ("Warm Network", "Founder/operator contact", "US/UK/NG/Remote", "", "", "", "Known trust or referral path", "", "Referral Ask", "Not Contacted", 0, 0.1, "=L5*M5", "Send direct ask", date(2026, 7, 12), "", ""),
        ("Swim/Sports Niche", "Swim school", "US/UK/Canada/Australia", "", "", "", "Trial, assessment, class placement, waitlist, inquiry form, waiver, payment, or follow-up clue", "", "Paid Audit", "Not Contacted", 400, 0.1, "=L6*M6", "Find owner/operator and personalize", date(2026, 7, 12), "", ""),
        ("Swim/Sports Niche", "Adult swim / triathlon coach", "US/UK/Canada/Australia", "", "", "", "Booking, video analysis, assessment, membership, recurring follow-up, or lead capture clue", "", "Paid Audit", "Not Contacted", 400, 0.1, "=L7*M7", "Personalize lead-to-enrollment message", date(2026, 7, 12), "", ""),
        ("Agency/Operator", "Small agency", "US/UK/Canada/Australia", "", "", "", "Client onboarding, reporting, recurring status, intake docs, dashboard, or follow-up pain", "", "Implementation Sprint", "Not Contacted", 1800, 0.1, "=L8*M8", "Send agency audit offer", date(2026, 7, 13), "", ""),
        ("Corporate Wellness", "HR/People/Ops contact", "Nigeria/remote", "", "", "", "Wellness, team-building, staff health, beginner swimming, safety, or community fit", "", "Corporate Intro", "Not Contacted", 500, 0.1, "=L9*M9", "Ask for intro", date(2026, 7, 15), "", ""),
    ]
    start_row = 5
    for idx, row in enumerate(segment_plan, start=start_row):
        pipe.cell(idx, 1, date(2026, 7, 12))
        for col, value in enumerate(row, start=2):
            pipe.cell(idx, col, value)
        pipe.cell(idx, 1).number_format = "yyyy-mm-dd"
        pipe.cell(idx, 16).number_format = "yyyy-mm-dd"
    for idx in range(start_row + len(segment_plan), start_row + 60):
        pipe.cell(idx, 1, "")
        pipe.cell(idx, 14, f"=L{idx}*M{idx}")
    style_table(pipe, 4, start_row + 59, 1, len(p_headers))
    add_list_validation(pipe, f"B5:B{start_row + 59}", tracks_list)
    add_list_validation(pipe, f"J5:J{start_row + 59}", offers)
    add_list_validation(pipe, f"K5:K{start_row + 59}", stages)
    set_widths(
        pipe,
        {
            "A": 13,
            "B": 20,
            "C": 24,
            "D": 22,
            "E": 28,
            "F": 22,
            "G": 38,
            "H": 56,
            "I": 22,
            "J": 22,
            "K": 18,
            "L": 12,
            "M": 12,
            "N": 15,
            "O": 32,
            "P": 16,
            "Q": 16,
            "R": 36,
        },
    )
    pipe.freeze_panes = "A5"
    pipe.auto_filter.ref = f"A4:R{start_row + 59}"

    # Prospect Criteria
    crit = wb.create_sheet("Prospect Criteria")
    style_title(crit, "Prospect Criteria", "Use this to decide who belongs in the first 40 prospects.", 8)
    c_headers = ["Quota", "Segment", "Ideal Buyer", "Pain Clues", "Where To Find", "First Offer", "Disqualifiers", "July Priority"]
    for c_idx, header in enumerate(c_headers, start=1):
        crit.cell(4, c_idx, header)
    style_header_row(crit, 4, 1, len(c_headers), COLORS["teal"])
    criteria = [
        (10, "Warm founder/operator contacts", "People who already know your ability or can introduce you to operators.", "Manual admin, missed follow-ups, data scattered across tools, repeated reporting.", "LinkedIn, WhatsApp, former projects, founders in your circle.", "Referral ask or paid workflow audit.", "No trust path, no operational pain, no budget owner.", "Highest"),
        (15, "Western swim schools / sports academies", "Owners or operators of swim schools, adult swim programs, triathlon swim coaches, kids sports academies.", "Inquiry form, trial/assessment booking, class levels, waitlist, waivers, multiple locations, payment follow-up.", "USA Swimming, USMS, Google Maps, local directories, LinkedIn.", "Lead-to-Enrollment Leak Audit.", "No clear owner/contact, no booking/enrollment flow, tiny volunteer-only club.", "Market test"),
        (10, "Agencies / small operators", "Small agencies, implementation shops, consultants, SaaS operators, or service teams.", "Client onboarding, intake docs, reporting, dashboards, recurring status updates, messy handoff.", "LinkedIn, founder communities, agency directories, warm intros.", "Client Onboarding/Reporting Workflow Audit.", "Pure content agencies with no ops pain, no decision maker, no budget.", "High"),
        (5, "Backup local/corporate SwimBuddz", "HR/People/Ops teams, community leaders, wellness decision makers, clubs.", "Employee wellness, beginner adult swimming, safety, community, team bonding.", "Warm network, LinkedIn, local companies, past SwimBuddz contacts.", "Corporate swim/wellness intro call.", "No warm path, no wellness budget, long procurement.", "Backup"),
    ]
    for r_idx, row in enumerate(criteria, start=5):
        for c_idx, value in enumerate(row, start=1):
            crit.cell(r_idx, c_idx, value)
    style_table(crit, 4, 4 + len(criteria), 1, len(c_headers))
    set_widths(crit, {"A": 10, "B": 30, "C": 44, "D": 56, "E": 42, "F": 36, "G": 42, "H": 16})

    # Offers and Templates
    offers_ws = wb.create_sheet("Offers & Templates")
    style_title(offers_ws, "Offers & Templates", "Use direct, concrete offers. Do not sell generic AI automation.", 5)
    o_headers = ["Track", "Asset", "Copy / Script", "Use When", "Notes"]
    for c_idx, header in enumerate(o_headers, start=1):
        offers_ws.cell(4, c_idx, header)
    style_header_row(offers_ws, 4, 1, len(o_headers), COLORS["blue"])
    templates = [
        (
            "Strategy",
            "July rule",
            "By July 31, collect $2k from either one $1.5k-$2.5k implementation sprint or 4-5 $300-$500 paid audits. Primary cash path is warm-network tech/ops gigs. Swim/sports academy is a 5-day niche test, not the only bet.",
            "Dashboard / daily reminder",
            "This prevents overcommitting to a slow cold market.",
        ),
        (
            "Warm Network",
            "Direct ask",
            "Hey [Name], quick ask. I am looking to close 1-2 fixed-scope workflow automation projects before July 31. I am strongest at AI/backend systems, integrations, internal tools, dashboards, and operational workflows. Do you know any founder/operator losing time or leads because a manual process is messy? I can do either a 48h paid audit or a 10-day implementation sprint.",
            "People who know your work",
            "Highest short-term probability because trust already exists.",
        ),
        (
            "Swim/Sports Niche",
            "Cold opener",
            "Hi [Name], I run SwimBuddz and noticed [specific clue from their booking/enrollment flow]. A lot of swim schools lose leads between inquiry, assessment, trial, class placement, payment, and follow-up. I am testing a 48h Lead-to-Enrollment Leak Audit: Loom teardown, flow map, quick fixes, and an implementation plan. Would it be useful if I sent one specific observation from your current flow?",
            "Owners/operators with visible lead flow",
            "Ask permission first if there is no warm connection.",
        ),
        (
            "Agency/Operator",
            "Cold opener",
            "Hi [Name], I noticed your team handles [client/service type]. Agencies often lose time in client onboarding and reporting: intake docs, status updates, dashboards, recurring reports, and follow-ups. I build fixed-scope backend/automation workflows. Open to me sending one specific observation about a workflow you could tighten?",
            "Small agencies/operators",
            "This is the fastest adjacent niche if swim replies are weak.",
        ),
        (
            "Closing",
            "Paid audit close",
            "Based on what you described, I would not start with a full build. I would do a 48h workflow audit first: map the current flow, identify leaks, propose quick wins, and give you the implementation plan. It is $400 and I credit it toward the sprint if we build. Want me to do that this week?",
            "After a call or strong reply",
            "Do not leave conversations at 'let me know'.",
        ),
        (
            "Closing",
            "Sprint close",
            "The sprint would focus on one workflow end-to-end: [workflow]. Deliverables: intake/booking flow, integrations, reminders, payment/status tracking, simple dashboard, error handling, and handover notes. Timeline: 10 business days. Price: $1.5k-$2.5k depending on scope. To reserve the week, I take a 50 percent deposit.",
            "When audit/call reveals concrete pain",
            "Keep scope narrow; one workflow only.",
        ),
        (
            "Follow-up",
            "Follow-up 1",
            "Quick follow-up on this. Is lead follow-up / onboarding / reporting currently a real problem for you, or is it already handled well enough?",
            "2 days after no reply",
            "Forces useful market feedback.",
        ),
        (
            "Follow-up",
            "Referral follow-up",
            "No worries if not you. Is there one founder/operator you think this would be more relevant for?",
            "Warm network non-buyer",
            "Converts non-buyers into intros.",
        ),
    ]
    for r_idx, row in enumerate(templates, start=5):
        for c_idx, value in enumerate(row, start=1):
            offers_ws.cell(r_idx, c_idx, value)
    style_table(offers_ws, 4, 4 + len(templates), 1, len(o_headers))
    set_widths(offers_ws, {"A": 20, "B": 22, "C": 86, "D": 32, "E": 42})

    # Gate Logic
    gate = wb.create_sheet("Gate Logic")
    style_title(gate, "Gate Logic", "Use the market response to decide. Do not argue with weak signals.", 6)
    g_headers = ["Date", "Question", "Green Signal", "Red Signal", "Action If Green", "Action If Red"]
    for c_idx, header in enumerate(g_headers, start=1):
        gate.cell(4, c_idx, header)
    style_header_row(gate, 4, 1, len(g_headers), COLORS["orange"])
    gate_rows = [
        ("2026-07-16", "Is swim/sports academy a real July cash path?", "5 replies, 2 calls, or 1 paid-audit interest.", "Fewer than 3 replies and no call momentum.", "Keep swim at 40 percent of outreach and continue Loom teardowns.", "Reduce swim to 20 percent; shift to warm network and agency/operator."),
        ("2026-07-20", "Is any channel producing buyer conversations?", "1 proposal or $300+ paid audit sold.", "Only polite replies and no paid ask accepted.", "Push proposals and sprint deposits.", "Make direct high-trust asks; lower friction with $300 audit."),
        ("2026-07-24", "Is cash collection on track?", "$500+ collected or signed sprint.", "No committed money.", "Deliver fast and ask for referrals/deposit.", "Stop cold experiments; work only strongest 10 leads."),
        ("2026-07-28", "Can $2k close by Jul 31?", "$1k+ collected/committed and at least one sprint in negotiation.", "No committed buyer.", "Push sprint deposit and paid audit bundle.", "Use immediate cash alternatives: direct contract work, referrals, local corporate intro."),
    ]
    for r_idx, row in enumerate(gate_rows, start=5):
        for c_idx, value in enumerate(row, start=1):
            gate.cell(r_idx, c_idx, value)
    style_table(gate, 4, 4 + len(gate_rows), 1, len(g_headers))
    set_widths(gate, {"A": 14, "B": 38, "C": 36, "D": 36, "E": 42, "F": 48})

    # Reminder Config
    config = wb.create_sheet("Reminder Config")
    style_title(config, "Reminder Config", "Paste Apps Script Code into Extensions > Apps Script, then run syncPlanReminders.", 4)
    conf_rows = [
        ("Key", "Value", "Description"),
        ("Calendar ID", "primary", "Use primary calendar unless you paste a specific calendar ID."),
        ("Event Prefix", "Revenue Sprint", "Prefix added to calendar events."),
        ("Max Events Per Run", 20, "Keeps Apps Script under event creation limits. Run again later for remaining rows."),
        ("Dry Run", "FALSE", "Set TRUE to test logging without creating events."),
        ("Timezone", "Africa/Lagos", "Your local planning timezone."),
        ("Start Date", "2026-07-12", "First date in this workbook."),
        ("End Date", "2026-07-31", "Last date in this workbook."),
    ]
    for r_idx, row in enumerate(conf_rows, start=4):
        for c_idx, value in enumerate(row, start=1):
            config.cell(r_idx, c_idx, value)
    style_header_row(config, 4, 1, 3, COLORS["blue"])
    style_table(config, 4, 4 + len(conf_rows) - 1, 1, 3)
    set_widths(config, {"A": 24, "B": 28, "C": 72})

    # Apps Script Code
    code = wb.create_sheet("Apps Script Code")
    style_title(code, "Apps Script Code", "Copy the code below into Apps Script for idempotent calendar reminders.", 1)
    code.cell(4, 1, APP_SCRIPT)
    code.cell(4, 1).alignment = Alignment(wrap_text=True, vertical="top")
    code.column_dimensions["A"].width = 140
    code.row_dimensions[4].height = 720

    # Source Notes
    src = wb.create_sheet("Source Notes")
    style_title(src, "Source Notes", "Research basis for the strategy. These links should be rechecked before a larger campaign.", 5)
    s_headers = ["Source", "Link", "What It Supports", "Planning Implication", "Reliability"]
    for c_idx, header in enumerate(s_headers, start=1):
        src.cell(4, c_idx, header)
    style_header_row(src, 4, 1, len(s_headers), COLORS["teal"])
    sources = [
        ("USA Swimming", "https://www.usaswimming.org/find-a-team", "Large organized swim team ecosystem and directories.", "There are many swim-related organizations to target, but club/team is not the same as paid school/operator.", "Official"),
        ("U.S. Masters Swimming", "https://www.usms.org/clubs", "Adult swim club/program ecosystem.", "Adult swim and triathlon-related programs are a relevant adjacent audience.", "Official"),
        ("Jackrabbit Class", "https://www.jackrabbitclass.com/", "Youth activity businesses already buy registration, scheduling, enrollment, billing, and engagement software.", "Customers pay for operations tools, but there are strong incumbents. Sell implementation around the current stack, not a generic OS.", "Vendor"),
        ("iClassPro", "https://www.iclasspro.com/", "Class/camp registration, payments, customer portal, and activity-center management.", "Competition is real. The wedge must be a specific leak audit and implementation, not 'software for swim schools'.", "Vendor"),
        ("Uplifter", "https://www.uplifterinc.com/", "Sports/activity software for registration, payments, waitlists, communications, forms, and reporting.", "Lead-to-enrollment and admin workflows are real pains, but many buyers already have a platform.", "Vendor"),
        ("Amilia", "https://www.amilia.com/", "Recreation/YMCA software positioning around registration, membership, operations, and admin time savings.", "Operational time savings and enrollment capacity are business outcomes buyers understand.", "Vendor"),
        ("AgencyAnalytics", "https://agencyanalytics.com/", "Agencies buy reporting/dashboard automation.", "Agency onboarding/reporting is a credible adjacent niche and may close faster than cold swim schools.", "Vendor"),
    ]
    for r_idx, row in enumerate(sources, start=5):
        for c_idx, value in enumerate(row, start=1):
            src.cell(r_idx, c_idx, value)
    style_table(src, 4, 4 + len(sources), 1, len(s_headers))
    set_widths(src, {"A": 22, "B": 48, "C": 54, "D": 66, "E": 14})

    # Lookups
    lookups = wb.create_sheet("Lookups")
    style_title(lookups, "Lookups", "Dropdown source values.", 5)
    lookup_cols = {
        "A": ("Statuses", ["Not Started", "In Progress", "Waiting", "Done", "Skipped"]),
        "B": ("Tracks", ["Warm Network", "Agency/Operator", "Swim/Sports Niche", "Corporate Wellness", "StrokeLab", "Admin/Review", "Strategy", "Proof", "Closing", "Delivery/Closing", "Other"]),
        "C": ("Stages", ["Not Contacted", "Messaged", "Replied", "Call Booked", "Call Held", "Proposal Sent", "Paid Audit", "Sprint Won", "Lost", "Nurture"]),
        "D": ("Priorities", ["P0", "P1", "P2"]),
        "E": ("Offers", ["Paid Audit", "Implementation Sprint", "Referral Ask", "Corporate Intro", "Digital Product", "Other"]),
    }
    for col, (header, values) in lookup_cols.items():
        lookups[f"{col}4"] = header
        for idx, value in enumerate(values, start=5):
            lookups[f"{col}{idx}"] = value
    style_header_row(lookups, 4, 1, 5, COLORS["gray"])
    style_table(lookups, 4, 15, 1, 5)
    set_widths(lookups, {"A": 18, "B": 24, "C": 20, "D": 14, "E": 24})

    for sheet in wb.worksheets:
        sheet.sheet_view.showGridLines = False

    wb.save(OUT_FILE)
    return OUT_FILE


if __name__ == "__main__":
    path = build_workbook()
    wb = load_workbook(path, data_only=False)
    print(path)
    for ws in wb.worksheets:
        print(f"{ws.title}: {ws.max_row} rows x {ws.max_column} cols")

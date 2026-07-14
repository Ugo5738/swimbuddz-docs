from datetime import date
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


WORKBOOK = Path("outputs/july_2026_revenue_sprint_restrategy/daniel_july_12_31_2026_revenue_restrategy.xlsx")

COLORS = {
    "navy": "0F172A",
    "blue": "0E7490",
    "teal": "0F766E",
    "light_blue": "D9EEF7",
    "white": "FFFFFF",
}

thin = Side(style="thin", color="CBD5E1")
border = Border(left=thin, right=thin, top=thin, bottom=thin)


def style_table(ws, min_row, max_row, min_col, max_col):
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    for r in range(min_row + 1, max_row + 1):
        fill = COLORS["white"] if r % 2 else COLORS["light_blue"]
        for c in range(min_col, max_col + 1):
            ws.cell(r, c).fill = PatternFill("solid", fgColor=fill)


def style_header(ws, row, end_col, fill=COLORS["blue"]):
    for col in range(1, end_col + 1):
        cell = ws.cell(row, col)
        cell.font = Font(bold=True, color=COLORS["white"])
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = border


def prospects():
    warm = [
        ("Warm Network", "Founder/operator contact", "Your network", "Warm Contact 1", "", "", "Knows your work or can intro you to founders/operators.", "Referral Ask", 0, 0.20, "Replace with actual name and send direct ask."),
        ("Warm Network", "Former colleague/client", "Your network", "Warm Contact 2", "", "", "Has seen your technical ability or execution.", "Referral Ask", 0, 0.20, "Replace with actual name and send direct ask."),
        ("Warm Network", "Founder friend", "Your network", "Warm Contact 3", "", "", "Can name one founder/operator with manual workflow pain.", "Referral Ask", 0, 0.20, "Replace with actual name and send direct ask."),
        ("Warm Network", "SwimBuddz member/contact", "Your network", "Warm Contact 4", "", "", "Trust path through SwimBuddz/community.", "Referral Ask", 0, 0.15, "Replace with actual name and ask for intro."),
        ("Warm Network", "Open-source/tech contact", "Your network", "Warm Contact 5", "", "", "Can refer agency/founder/operator contacts.", "Referral Ask", 0, 0.15, "Replace with actual name and ask for intro."),
        ("Warm Network", "Previous school/Unilag contact", "Your network", "Warm Contact 6", "", "", "AI/ML credibility path.", "Referral Ask", 0, 0.15, "Replace with actual name and ask for intro."),
        ("Warm Network", "Rolf/expert path", "Your network", "Warm Contact 7", "", "", "Can advise or introduce higher-quality buyers.", "Referral Ask", 0, 0.15, "Replace with actual name and ask for intro."),
        ("Warm Network", "Community founder", "Your network", "Warm Contact 8", "", "", "Likely knows people running small operations.", "Referral Ask", 0, 0.15, "Replace with actual name and ask for intro."),
        ("Warm Network", "Past project stakeholder", "Your network", "Warm Contact 9", "", "", "Can validate backend/automation execution.", "Referral Ask", 0, 0.15, "Replace with actual name and ask for intro."),
        ("Warm Network", "Trusted friend with founder network", "Your network", "Warm Contact 10", "", "", "Can connect you quickly without cold trust problem.", "Referral Ask", 0, 0.15, "Replace with actual name and ask for intro."),
    ]

    swim = [
        ("Swim/Sports Niche", "Swim school", "US", "Goldfish Swim School", "", "https://goldfishswimschool.com/", "Assessment, registration portal, class levels, multiple locations, make-up lessons.", "Paid Audit", 400, 0.10, "Target a local franchise owner/operator, not only HQ."),
        ("Swim/Sports Niche", "Swim school", "US", "British Swim School", "", "https://britishswimschool.com/", "Location finder, swim assessment, programs for children/adults/adaptive aquatics.", "Paid Audit", 400, 0.10, "Target a local owner/operator."),
        ("Swim/Sports Niche", "Swim school", "US", "Aqua-Tots Swim Schools", "", "https://www.aqua-tots.com/", "Enrollment, levels, franchise locations, likely parent follow-up workflow.", "Paid Audit", 400, 0.10, "Target regional/franchise operator."),
        ("Swim/Sports Niche", "Swim school", "US", "SafeSplash Swim School", "", "https://www.safesplash.com/", "Kids/adult lessons, locations, booking/enrollment flow.", "Paid Audit", 400, 0.10, "Target location-level operator."),
        ("Swim/Sports Niche", "Swim school", "US", "Big Blue Swim School", "", "https://bigblueswimschool.com/", "Class enrollment, parent-facing scheduling, multi-location operation.", "Paid Audit", 400, 0.08, "Likely has mature systems; use as research or local franchise target."),
        ("Swim/Sports Niche", "Swim school", "US", "Emler Swim School", "", "https://emlerswimschool.com/", "Class levels, enrollment, parent/student progression.", "Paid Audit", 400, 0.08, "Look for individual location decision maker."),
        ("Swim/Sports Niche", "Swim school", "US", "Foss Swim School", "", "https://www.fossswimschool.com/", "Enrollment, class levels, multi-location operations.", "Paid Audit", 400, 0.08, "Research local operator before messaging."),
        ("Swim/Sports Niche", "Swim school", "US", "Bear Paddle Swim School", "", "https://bearpaddle.com/", "Lesson enrollment and progression flow.", "Paid Audit", 400, 0.08, "Research local/franchise contact."),
        ("Swim/Sports Niche", "Private swim lessons", "US/Canada", "Sunsational Swim School", "", "https://www.sunsationalswimschool.com/", "Private lesson inquiry, matching, scheduling, instructor/customer coordination.", "Paid Audit", 400, 0.10, "Lead matching workflow may be a strong angle."),
        ("Swim/Sports Niche", "Private swim lessons", "US/Canada", "AquaMobile Swim School", "", "https://aquamobileswim.com/", "At-home lesson booking, instructor matching, payment/follow-up flow.", "Paid Audit", 400, 0.10, "Lead-to-instructor matching workflow may be relevant."),
        ("Swim/Sports Niche", "Swim analysis/coaching", "US", "SwimLabs", "", "https://www.swimlabs.com/", "Video analysis, lessons, locations, assessment/booking workflow.", "Paid Audit", 400, 0.08, "Good StrokeLab proof comparison."),
        ("Swim/Sports Niche", "Swim school", "Australia", "Carlile Swimming", "", "https://www.carlile.com.au/", "Learn-to-swim program, levels, enrollment, multiple centers.", "Paid Audit", 400, 0.08, "May be too mature; use as research or local contact."),
        ("Swim/Sports Niche", "Swim school", "Australia", "State Swim", "", "https://www.stateswim.com.au/", "Class levels, booking/enrollment, multiple locations.", "Paid Audit", 400, 0.08, "Research franchise/location operators."),
        ("Swim/Sports Niche", "Swim school", "Australia", "JUMP! Swim Schools", "", "https://jumpswimschools.com.au/", "Small-class swim school model, franchise locations, enrollment flow.", "Paid Audit", 400, 0.08, "Target local operator if contact is visible."),
        ("Swim/Sports Niche", "Swim school", "Australia", "Kingswim", "", "https://www.kingswim.com.au/", "Lessons, class levels, customer booking, multi-location operations.", "Paid Audit", 400, 0.08, "Research individual center operators."),
    ]

    agency = [
        ("Agency/Operator", "B2B SaaS agency", "US/Canada", "Powered by Search", "", "https://www.poweredbysearch.com/", "Client onboarding, reporting, campaign dashboards, recurring delivery process.", "Implementation Sprint", 1800, 0.08, "Target ops/delivery leader, not generic contact."),
        ("Agency/Operator", "Inbound agency", "US", "SmartBug Media", "", "https://www.smartbugmedia.com/", "Client reporting, onboarding, HubSpot/RevOps workflows.", "Implementation Sprint", 1800, 0.08, "May already be mature; use specific workflow angle."),
        ("Agency/Operator", "Performance agency", "US", "KlientBoost", "", "https://klientboost.com/", "Client reporting and recurring performance communication.", "Implementation Sprint", 1800, 0.07, "Likely mature; prioritize if you can find ops pain."),
        ("Agency/Operator", "SaaS marketing agency", "US/UK", "Directive Consulting", "", "https://directiveconsulting.com/", "SaaS client onboarding, reporting, dashboards.", "Implementation Sprint", 1800, 0.07, "Likely mature; use as research if no contact path."),
        ("Agency/Operator", "Content agency", "US", "Siege Media", "", "https://www.siegemedia.com/", "Content production workflow, client approvals, reporting.", "Implementation Sprint", 1800, 0.08, "Workflow operations may be more relevant than AI pitch."),
        ("Agency/Operator", "SEO agency", "US", "SimpleTiger", "", "https://www.simpletiger.com/", "SEO deliverables, reporting, recurring client workflows.", "Implementation Sprint", 1800, 0.08, "Look for founder/ops contact."),
        ("Agency/Operator", "Growth agency", "US", "NoGood", "", "https://nogood.io/", "Growth experiments, reporting, internal workflow handoffs.", "Implementation Sprint", 1800, 0.07, "May be too sophisticated; use if specific pain clue exists."),
        ("Agency/Operator", "Growth agency", "US", "Tuff Growth", "", "https://tuffgrowth.com/", "Client onboarding/reporting and marketing ops workflows.", "Implementation Sprint", 1800, 0.08, "Find founder/ops contact."),
        ("Agency/Operator", "Content agency", "US", "Grow and Convert", "", "https://www.growandconvert.com/", "Content workflow, research, client reporting, handoff process.", "Implementation Sprint", 1800, 0.08, "Specific workflow observation needed."),
        ("Agency/Operator", "Content agency", "US", "Codeless", "", "https://getcodeless.com/", "Content operations, approvals, recurring reporting.", "Implementation Sprint", 1800, 0.08, "Specific workflow observation needed."),
    ]

    corporate = [
        ("Corporate Wellness", "Fintech/company", "Nigeria", "Flutterwave", "", "https://flutterwave.com/", "Corporate wellness/community angle; only pursue through warm intro.", "Corporate Intro", 500, 0.05, "Do not cold pitch HR first; find warm path."),
        ("Corporate Wellness", "Fintech/company", "Nigeria", "Paystack", "", "https://paystack.com/", "Corporate wellness/community angle; only pursue through warm intro.", "Corporate Intro", 500, 0.05, "Do not cold pitch HR first; find warm path."),
        ("Corporate Wellness", "Fintech/company", "Nigeria", "PiggyVest", "", "https://www.piggyvest.com/", "Corporate wellness/community angle; only pursue through warm intro.", "Corporate Intro", 500, 0.05, "Do not cold pitch HR first; find warm path."),
        ("Corporate Wellness", "Tech talent/company", "Remote/Africa", "Andela", "", "https://andela.com/", "Wellness/community/team-building angle; only pursue through warm intro.", "Corporate Intro", 500, 0.05, "Find people/ops contact through network."),
        ("Corporate Wellness", "Fintech/company", "Nigeria", "Cowrywise", "", "https://cowrywise.com/", "Corporate wellness/community angle; only pursue through warm intro.", "Corporate Intro", 500, 0.05, "Find warm intro first."),
    ]

    return warm + swim + agency + corporate


def main():
    wb = load_workbook(WORKBOOK)

    if "Starter Prospects" in wb.sheetnames:
        del wb["Starter Prospects"]
    ws = wb.create_sheet("Starter Prospects", 4)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)
    ws.cell(1, 1, "Starter 40 Prospects / Slots")
    ws.cell(1, 1).font = Font(bold=True, size=18, color=COLORS["white"])
    ws.cell(1, 1).fill = PatternFill("solid", fgColor=COLORS["navy"])
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=12)
    ws.cell(2, 1, "Replace the 10 warm placeholders with real people. Public cold prospects still need owner/operator contact research before messaging.")
    ws.cell(2, 1).font = Font(italic=True, color="475569")

    headers = [
        "Track",
        "Segment",
        "Market",
        "Prospect",
        "Contact",
        "URL / LinkedIn",
        "Pain Clue",
        "Offer",
        "Amount",
        "Probability",
        "Next Action",
        "Stage",
    ]
    for col, header in enumerate(headers, start=1):
        ws.cell(4, col, header)
    style_header(ws, 4, len(headers), COLORS["teal"])

    for row_idx, row in enumerate(prospects(), start=5):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row_idx, col_idx, value)
        ws.cell(row_idx, 12, "Not Contacted")

    style_table(ws, 4, 4 + len(prospects()), 1, len(headers))
    widths = {
        "A": 20,
        "B": 24,
        "C": 20,
        "D": 32,
        "E": 22,
        "F": 40,
        "G": 58,
        "H": 22,
        "I": 12,
        "J": 12,
        "K": 44,
        "L": 18,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:L{4 + len(prospects())}"

    pipe = wb["Pipeline Tracker"]
    pipe_headers = {pipe.cell(4, col).value: col for col in range(1, pipe.max_column + 1)}
    for idx, row in enumerate(prospects(), start=5):
        track, segment, market, prospect, contact, url, pain, offer, amount, prob, next_action = row
        pipe.cell(idx, pipe_headers["Date Added"], date(2026, 7, 12))
        pipe.cell(idx, pipe_headers["Track"], track)
        pipe.cell(idx, pipe_headers["Segment"], segment)
        pipe.cell(idx, pipe_headers["Market"], market)
        pipe.cell(idx, pipe_headers["Prospect"], prospect)
        pipe.cell(idx, pipe_headers["Contact"], contact)
        pipe.cell(idx, pipe_headers["URL / LinkedIn"], url)
        pipe.cell(idx, pipe_headers["Pain Clue"], pain)
        pipe.cell(idx, pipe_headers["Offer"], offer)
        pipe.cell(idx, pipe_headers["Stage"], "Not Contacted")
        pipe.cell(idx, pipe_headers["Amount"], amount)
        pipe.cell(idx, pipe_headers["Probability"], prob)
        pipe.cell(idx, pipe_headers["Weighted Value"], f"=L{idx}*M{idx}")
        pipe.cell(idx, pipe_headers["Next Action"], next_action)
        pipe.cell(idx, pipe_headers["Next Action Date"], date(2026, 7, 12))
        pipe.cell(idx, pipe_headers["Notes"], "Starter seed. Verify contact before sending.")

    wb.save(WORKBOOK)
    print(WORKBOOK)
    print(f"Seeded {len(prospects())} starter prospect rows.")


if __name__ == "__main__":
    main()
